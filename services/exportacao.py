import os
import csv
import shutil
import datetime

BASE_DIR   = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
BACKUP_DIR = os.path.join(BASE_DIR, "backups")
os.makedirs(BACKUP_DIR, exist_ok=True)

CABECALHOS = ["ID", "Nome", "Email", "Telefone", "Sexo",
              "Data Nascimento", "Endereço", "Curso"]


def exportar_excel(dados, path):
    
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Alunos"

    h_fill = PatternFill("solid", fgColor="042C53")
    h_font = Font(color="FFFFFF", bold=True, size=10)
    borda  = Border(**{s: Side(style="thin", color="D3D1C7")
                       for s in ("left", "right", "top", "bottom")})

    for c, h in enumerate(CABECALHOS, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill, cell.font, cell.border = h_fill, h_font, borda
        cell.alignment = Alignment(horizontal="center", vertical="center")

    alt = PatternFill("solid", fgColor="F1EFE8")
    for r, row in enumerate(dados, 2):
        fill = alt if r % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for c, val in enumerate(row[:8], 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.fill, cell.border = fill, borda
            cell.alignment = Alignment(vertical="center")

    for i, w in enumerate([6, 30, 30, 16, 8, 16, 35, 25], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 22
    wb.save(path)


def exportar_csv(dados, path):
    
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(CABECALHOS)
        for row in dados:
            w.writerow(row[:8])


def fazer_backup():
    
    db_path = os.path.join(BASE_DIR, "estudante.db")
    if not os.path.exists(db_path):
        raise FileNotFoundError("Banco de dados não encontrado.")
    ts      = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    destino = os.path.join(BACKUP_DIR, f"estudante_backup_{ts}.db")
    shutil.copy2(db_path, destino)
    return destino


def importar_csv(path):
    
    registros = []
    erros     = []

    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        linhas = list(reader)

    if not linhas:
        return [], ["Arquivo vazio."]

    
    inicio = 1 if linhas[0][0].strip().upper() in ("ID", "NOME", "NAME") else 0

    for i, linha in enumerate(linhas[inicio:], start=inicio + 1):
        
        cols = [c.strip() for c in linha]

        
        if len(cols) == 9:
            cols = cols[1:]

        if len(cols) < 7:
            erros.append(f"Linha {i}: colunas insuficientes ({len(cols)} encontradas).")
            continue

        nome, email, telefone, sexo, data, endereco, curso = cols[:7]
        foto = cols[7] if len(cols) > 7 else ""

        if not nome:
            erros.append(f"Linha {i}: nome vazio, ignorada.")
            continue

        registros.append([nome, email, telefone, sexo, data, endereco, curso, foto])

    return registros, erros


def importar_excel(path):
    
    import openpyxl
    registros = []
    erros     = []

    wb = openpyxl.load_workbook(path)
    ws = wb.active
    linhas = list(ws.iter_rows(values_only=True))

    if not linhas:
        return [], ["Arquivo vazio."]

    
    inicio = 1 if str(linhas[0][0]).strip().upper() in ("ID", "NOME", "NAME") else 0

    for i, linha in enumerate(linhas[inicio:], start=inicio + 1):
        cols = [str(c).strip() if c is not None else "" for c in linha]

        
        if len(cols) == 9:
            cols = cols[1:]

        if len(cols) < 7:
            erros.append(f"Linha {i}: colunas insuficientes.")
            continue

        nome, email, telefone, sexo, data, endereco, curso = cols[:7]
        foto = cols[7] if len(cols) > 7 else ""

        if not nome:
            erros.append(f"Linha {i}: nome vazio, ignorada.")
            continue

        registros.append([nome, email, telefone, sexo, data, endereco, curso, foto])

    return registros, erros